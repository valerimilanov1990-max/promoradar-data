#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Промо Радар — дневен събирач на данни v3 (върви на GitHub Actions).

Какво е ново спрямо v2:
  * eBag минава през истински браузър (Playwright), а не през requests —
    така се хващат и цени, които се дорисуват с JavaScript.
  * Всеки продукт получава КОЛИЧЕСТВО и ЦЕНА ЗА ЕДИНИЦА (лв/кг, лв/л,
    лв/бр). Това е числото, по което цените реално се сравняват.
  * Диагностика: „stats“ казва колко неща е върнал всеки източник,
    а „errors“ — какво точно се е счупило. Никога празен изход без обяснение.
  * Всички ключове съществуват винаги (ebag, offers, basics, brochures,
    errors, stats) — приложението никога не получава непълен JSON.

Изход: data.json
"""
import csv
import datetime
import io
import json
import re
import traceback
import zipfile

import requests

UA = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/126.0 Safari/537.36",
      "Accept-Language": "bg-BG,bg;q=0.9"}

OUT = {
    "updated": datetime.datetime.now().strftime("%d.%m.%Y %H:%M"),
    "version": 3,
    "ebag": [],
    "offers": [],
    "basics": [],
    "community": [],
    "history": {},
    "brochures": [],
    "errors": [],
    "stats": {},
}

CFG = json.load(open("config.json", encoding="utf-8"))


def log_err(section, e):
    OUT["errors"].append(f"{section}: {type(e).__name__}: {e}")
    traceback.print_exc()


def note(section, msg):
    OUT["errors"].append(f"{section}: {msg}")


PRICE_RE = re.compile(r"(\d{1,4}[.,]\d{2})\s*(?:лв|lv)", re.IGNORECASE)

# ---------------------------------------------------------------------------
# Количество и цена за единица
# ---------------------------------------------------------------------------
# Редът има значение: по-дългите мерки (кг, мг, мл) СТОЯТ ПРЕДИ късите (г, л),
# иначе „200мг“ ще се разчете като „200 г“.
_UNITS = r"кг|kg|мг|mg|мл|ml|гр|г|g|литра|литър|л|l|броя|бройки|бр\.?|броя|pcs|db"

_MULTI_RE = re.compile(
    r"(\d{1,3})\s*[xх×]\s*(\d+(?:[.,]\d+)?)\s*(" + _UNITS + r")(?![а-яa-z])",
    re.IGNORECASE)
_SINGLE_RE = re.compile(
    r"(?<![\d.,])(\d+(?:[.,]\d+)?)\s*(" + _UNITS + r")(?![а-яa-z])",
    re.IGNORECASE)


def _to_base(amount, unit):
    """Връща (количество_в_базова_мярка, етикет) или None, ако мярката се пропуска."""
    u = unit.lower().rstrip(".")
    if u in ("кг", "kg"):
        return amount, "кг"
    if u in ("г", "гр", "g"):
        return amount / 1000.0, "кг"
    if u in ("мг", "mg"):
        return None                      # милиграми: не са хранителна мярка
    if u in ("л", "l", "литра", "литър"):
        return amount, "л"
    if u in ("мл", "ml"):
        return amount / 1000.0, "л"
    if u in ("бр", "бр.", "броя", "бройки", "pcs", "db"):
        return amount, "бр"
    return None


def parse_qty(name):
    """Изважда количеството от името на продукта.

    „Мляко прясно 3% 1л“        -> (1.0, 'л')
    „Кисело мляко 400 г“        -> (0.4, 'кг')
    „Вода Девин 6 х 1.5 л“      -> (9.0, 'л')
    „Яйца M 10 бр.“             -> (10.0, 'бр')
    Връща None, когато няма разпознаваемо количество.
    """
    if not name:
        return None
    txt = name.replace(" ", " ")

    m = _MULTI_RE.search(txt)
    if m:
        try:
            packs = float(m.group(1).replace(",", "."))
            each = float(m.group(2).replace(",", "."))
            base = _to_base(each, m.group(3))
            if base and 0 < packs <= 100:
                return round(base[0] * packs, 4), base[1]
        except ValueError:
            pass

    best = None
    for m in _SINGLE_RE.finditer(txt):
        try:
            val = float(m.group(1).replace(",", "."))
        except ValueError:
            continue
        base = _to_base(val, m.group(2))
        if not base:
            continue
        amount, label = base
        if not (0 < amount <= 100):
            continue
        # при няколко съвпадения взимаме последното (обикновено то е грамажът)
        best = (round(amount, 4), label)
    return best


def unit_price(price, name):
    """Връща (цена_за_единица, етикет) напр. (4.98, 'лв/кг') или None."""
    q = parse_qty(name)
    if not q or not price:
        return None
    amount, label = q
    if amount <= 0:
        return None
    up = price / amount
    if not (0 < up < 10000):
        return None
    # +1e-9 компенсира двоичната неточност: 1.99/0.4 е 4.9749999… във float,
    # което иначе се закръгля надолу до 4.97 вместо до 4.98.
    return round(up + 1e-9, 2), f"лв/{label}"


EUR_RATE = 1.95583   # фиксираният курс лев/евро


def collapse_currency(prices, curs=None, pct=None):
    """Маха дублиращите се цени, изписани в другата валута.

    След приемането на еврото магазините показват ЕДНА цена два пъти:
    „29.99 €“ и „58.66 лв“. Ако ги подадем нататък като две числа, кодът
    ги взима за цена и стара цена и обявява фалшиво намаление от 49%.
    Точно това се случи с офертите на Кауфланд.

    Разпознаваме двойките по съотношението 1.9558 и пазим левовата.
    """
    if not prices:
        return []
    curs = list(curs or [""] * len(prices))
    if len(curs) < len(prices):
        curs += [""] * (len(prices) - len(curs))

    # Ако на плочката пише намаление около 49%, то може и да е истинско —
    # тогава не пипаме нищо, за да не изтрием реална стара цена.
    if pct and 46 <= pct <= 52:
        return list(prices)

    pairs = list(zip(prices, curs))
    kept = []
    for v, c in pairs:
        if v <= 0:
            continue
        twin = any(o > v and abs(o / v - EUR_RATE) < 0.02 * EUR_RATE
                   for o, _ in pairs)
        if twin and c != "bgn":
            continue          # v е същата цена, но в евро — махаме я
        kept.append((v, c))

    if not kept:
        return list(prices)
    # Ако накрая всичко е в евро, превръщаме в лева
    if all(c == "eur" for _, c in kept):
        return [round(v * EUR_RATE, 2) for v, _ in kept]
    return [v for v, _ in kept]


def pick_price_pair(prices, pct=None):
    """Избира (текуща цена, стара цена) измежду намерените на плочката числа.

    Плочките често носят повече от две числа — цена за килограм, цена на
    вноска, тегло. Ако вземем сляпо най-малкото и най-голямото, старата
    цена излиза грешна.

    Затова: когато на плочката пише процент намаление, търсим двойката,
    която му отговаря най-точно. Иначе взимаме най-ниската цена и
    най-близката над нея.
    """
    vals = sorted({round(p, 2) for p in prices if 0.05 <= p <= 5000})
    if not vals:
        return None, None
    if len(vals) == 1:
        return vals[0], None

    if pct and 1 <= pct <= 95:
        target = pct / 100.0
        best, best_err = None, 1e9
        for i, p in enumerate(vals):
            for o in vals[i + 1:]:
                if o <= p:
                    continue
                err = abs((1 - p / o) - target)
                if err < best_err:
                    best_err, best = err, (p, o)
        # до 6 процентни пункта разлика приемаме за същата двойка
        if best and best_err <= 0.06:
            return best

    price = vals[0]
    # Старата цена е най-близката над текущата, но НЕ повече от 5 пъти:
    # съотношение 10 или 20 значи, че сме хванали цена за 100 г или за
    # кашон, а не истинско намаление от 95%.
    old = next((v for v in vals[1:] if price * 1.02 < v <= price * 5), None)
    return price, old


def enrich(item):
    """Добавя qty/unit/unitPrice към запис с полета name+price.

    Тук е и последната проверка срещу фалшиви намаления: ако „старата
    цена“ съвпада с цената за единица, тя не е стара цена, а цената за
    килограм/литър, попаднала в грешното поле. Проверката е числена, тоест
    хваща случая независимо как магазинът я е изписал.
    """
    name = item.get("name") or item.get("product", "")
    up = unit_price(item.get("price"), name)
    if up:
        item["unitPrice"], item["unitLabel"] = up
    q = parse_qty(name)
    if q:
        item["qty"], item["qtyUnit"] = q

    old = item.get("old")
    if old and up and abs(old - up[0]) <= max(0.02, up[0] * 0.01):
        item["old"] = None
        if item.get("pct"):
            item["pct"] = None
    return item


# Заглавия, които идват от рекламни рубрики, а не от стока. Улавят се тук,
# а не в JavaScript, защото същият текст стига и по други пътища.
JUNK_NAME_RE = re.compile(
    r"^\s*(тествай(те)?\b"
    r"|печеливш(и|ите)?\s+участниц"
    r"|тест\s+на\s+продукт"
    r"|играй(те)?\b|спечели\b|томбола\b"
    r"|виж(те)?\s+повече\b|научи\s+повече\b"
    r"|седмични\s+предложения?\b|наши(те)?\s+предложени"
    r"|нови\s+продукти\b|всички\s+продукти\b"
    r"|разгледай|каталог\b|брошура\b|листовк"
    r"|абонирай|бюлетин\b"
    # Открити при подреждането по категории: dm ги показва като плочки
    r"|информация\s+за\s+продукт|налично\s+за\s+доставка"
    r"|не\s+е\s+налично|кошница\s+с\s+грижа)", re.IGNORECASE)

# „25 пране (0,23 лв. за 1 пране)“ — това е ред с цена за единица, не име.
UNIT_AS_NAME_RE = re.compile(
    r"^\s*\d+([.,]\d+)?\s*(пранe|пране|бр|броя|г|гр|кг|мл|л|ml|g|kg|l)\b"
    r"|лв\.?\s*за\s+\d", re.IGNORECASE)

# Картинки, които не са стока: рейтинг-звездички, лога, иконки.
JUNK_IMG_RE = re.compile(
    r"review-ui|/stars?/|starfilled|staremp|rating|"
    r"/logo|placeholder|sprite|favicon|\.svg(\?|$)", re.IGNORECASE)


def is_junk_offer(o):
    """True = това не е стока и не бива да влиза във фийда."""
    name = (o.get("name") or "").strip()
    if len(name) < 3:
        return True
    if JUNK_NAME_RE.search(name) or UNIT_AS_NAME_RE.search(name):
        return True
    # само главни букви и препинателни знаци, без нито една буква — не е име
    if not re.search(r"[A-Za-zА-Яа-я]{3}", name):
        return True
    # „10 х 1 kg/опаковка“ — това е мярка, попаднала в полето за име
    if "/опаковка" in name:
        return True
    p = o.get("price")
    if not isinstance(p, (int, float)) or p <= 0 or p > 5000:
        return True
    return False


def clean_offer(o):
    """Маха подвеждащите полета, без да изхвърля целия запис."""
    if JUNK_IMG_RE.search(o.get("img") or ""):
        o["img"] = ""
    old, price = o.get("old"), o.get("price")
    # Стара цена, която е под или равна на новата, не е стара цена.
    if old is not None and price and old <= price * 1.005:
        o["old"] = None
        o.pop("pct", None)
    return o


# ---------------------------------------------------------------------------
# 1. eBag — през истински браузър (с резервен вариант requests)
# ---------------------------------------------------------------------------
def _ebag_from_html(pid, html):
    m_name = (re.search(r'og:title["\']?\s+content=["\']([^"\']+)', html)
              or re.search(r'content=["\']([^"\']+)["\']\s+property=["\']og:title', html)
              or re.search(r"<title>([^<]+)</title>", html))
    m_price = PRICE_RE.search(html)
    if not (m_name and m_price):
        return None
    name = re.sub(r"\s*[-|]\s*eBag\.bg\s*$", "", m_name.group(1)).strip()
    return {"id": str(pid), "name": name,
            "price": float(m_price.group(1).replace(",", "."))}


def scrape_ebag(ctx):
    ids = CFG.get("ebag_ids", [])
    base = CFG.get("ebag_base", "https://www.ebag.bg/x/")
    ok = 0
    for pid in ids:
        url = f"{base}{pid}"
        rec = None

        # 1) истински браузър — вижда и това, което JavaScript дорисува
        if ctx is not None:
            pg = None
            try:
                pg = ctx.new_page()
                pg.goto(url, wait_until="domcontentloaded", timeout=45000)
                pg.wait_for_timeout(2500)
                rec = _ebag_from_html(pid, pg.content())
                if not rec:
                    # цената понякога е само във видимия текст, не в мета
                    txt = pg.inner_text("body")
                    m = PRICE_RE.search(txt)
                    title = (pg.title() or f"Продукт {pid}")
                    if m:
                        rec = {"id": str(pid),
                               "name": re.sub(r"\s*[-|]\s*eBag\.bg\s*$", "", title).strip(),
                               "price": float(m.group(1).replace(",", "."))}
            except Exception as e:
                log_err(f"ebag/{pid}/browser", e)
            finally:
                if pg:
                    try:
                        pg.close()
                    except Exception:
                        pass

        # 2) резервно: обикновена заявка
        if not rec:
            try:
                r = requests.get(url, headers=UA, timeout=30)
                if r.status_code == 200:
                    rec = _ebag_from_html(pid, r.text)
                else:
                    note(f"ebag/{pid}", f"HTTP {r.status_code}")
            except Exception as e:
                log_err(f"ebag/{pid}/requests", e)

        if rec:
            OUT["ebag"].append(enrich(rec))
            ok += 1
        else:
            note(f"ebag/{pid}", "цената не се разпознава (нито в браузър, нито директно)")

    OUT["stats"]["ebag"] = f"{ok}/{len(ids)}"


# ---------------------------------------------------------------------------
# 2. Оферти от рендирания сайт (без OCR)
# ---------------------------------------------------------------------------
# --- ОБЩИ ПОМОЩНИЦИ ЗА ДВАТА МЕТОДА -----------------------------------
# Кауфланд мина по резервния метод, който нямаше поправките за имената и
# цените — затова логиката вече е ЕДНА и се вгражда и в двата.
JS_HELPERS = r"""
  const PRICE_G = /(\d{1,4}[.,]\d{2})/g;
  const PRICE_T = /\d{1,4}[.,]\d{2}/;

  // Редове като „цена за кг: 9.95“ са единични цени, не цената на продукта.
  // dm я пише като „(13,20 лв. за 1 L)“ — това е цена за литър, не стара
  // цена. Без този филтър всяка тяхна оферта излизаше с фалшиви -80%.
  const isUnitLine = (l) => /цена\s*за/i.test(l)
    || /\d\s*(лв|лева|eur|€)\.?\s*(\/|за)\s*\d*\s*(кг|kg|л\b|l\b|г\b|g\b|мл|ml|бр)/i.test(l)
    || /\/\s*(кг|kg|л\b|l\b|бр|100\s*(г|мл))/i.test(l)
    || /за\s+\d+([.,]\d+)?\s*(кг|kg|г\b|g\b|л\b|l\b|мл|ml|бр)\b/i.test(l);

  // След числото стои мярка -> това е ОБЕМ/ТЕГЛО, не цена („Мартини 0,75 л“).
  const UNIT_AFTER = /^\s*(л\b|l\b|кг|kg|гр|г\b|g\b|мл|ml|бр|броя|%|["“”]|x|х|см|cm|мм|mm|м\b|m\b|вт|w\b|kw|квт)/i;
  const CUR = /(лв|лева|bgn|€|eur)/i;
  const EURO = /€|eur/i;
  const LEVA = /лв|лева|bgn/i;

  // Връща [{v: число, c: 'eur'|'bgn'|''}] — валутата е нужна, защото
  // магазините изписват ЕДНАТА цена два пъти: в евро и в лева.
  const readPrices = (text, strict) => {
    const res = [];
    let m;
    PRICE_G.lastIndex = 0;
    while ((m = PRICE_G.exec(text)) !== null) {
      const end = m.index + m[0].length;
      const after = text.slice(end, end + 10);
      const before = text.slice(Math.max(0, m.index - 14), m.index);
      if (UNIT_AFTER.test(after)) continue;
      if (/\d$/.test(text.charAt(m.index - 1))) continue;
      const hasCur = CUR.test(after) || CUR.test(before);
      if (strict && !hasCur) continue;
      // Валутата се чете от ТЯСНО прозорче. Иначе символът от предишната
      // цена („16.61 €“) попада в контекста на следващата („32.49 лв“) и
      // двете излизат в евро — тогава кодът ги умножава по курса наново.
      const near = after.slice(0, 6);
      const pre  = before.slice(-3);
      let c = '';
      if (LEVA.test(near)) c = 'bgn';
      else if (EURO.test(near)) c = 'eur';
      else if (EURO.test(pre)) c = 'eur';       // „€40.39“
      else if (LEVA.test(pre)) c = 'bgn';
      res.push({v: parseFloat(m[1].replace(',', '.')), c: c});
    }
    return res;
  };

  const pricesOf = (el, wholeText) => {
    // ВАЖНО: четем от ЦЕЛИЯ текст на плочката, не от изолирания елемент с
    // клас „price“. Само там мярката стои непосредствено до числото и
    // „0,75 л“ се разпознава като обем, а не като цена от 75 стотинки.
    const lines = wholeText.split('\n').filter(l => !isUnitLine(l)).join('\n');
    let ms = readPrices(lines, true);      // първо с валутен знак наблизо
    if (!ms.length) ms = readPrices(lines, false);
    if (!ms.length) {
      // краен случай: плочката няма никакъв текст с цена наоколо
      const priceEls = el.querySelectorAll(
        '[class*="price" i],[class*="cena" i],[class*="Preis" i],[itemprop="price"]');
      let ptxt = '';
      priceEls.forEach(p => { ptxt += '\n' + (p.textContent || ''); });
      ms = readPrices(ptxt.split('\n').filter(l => !isUnitLine(l)).join('\n'), false);
    }
    return ms;
  };

  // alt текстове, които описват СНИМКАТА, а не продукта
  const ALT_PREFIX = /^\s*(изображение на|снимка на|картинка на|image of|photo of|picture of)\s*/i;
  const looksLikeDescription = (s) => {
    const words = s.split(/\s+/).filter(Boolean).length;
    if (words > 10) return true;
    if (/[.!?]$/.test(s.trim())) return true;
    return false;
  };

  // „rating.starFilled“, „product_title“ — вътрешни ключове на сайта,
  // изтекли в текста. Човешко име винаги има интервал или кирилица.
  const looksLikeCode = (s) => {
    if (/[А-Яа-я]/.test(s)) return false;
    if (/\s/.test(s)) return false;
    return /[._]/.test(s) || /^[a-z]+[A-Z]/.test(s);
  };

  // „0,2 L (13,20 лв. за 1 L)“ не е име на продукт, а ред с мерки.
  const mostlyNumeric = (s) => {
    const digits = (s.match(/\d/g) || []).length;
    const letters = (s.match(/[A-Za-zА-Яа-я]/g) || []).length;
    return letters < 5 || digits >= letters;
  };

  const nameOf = (el) => {
    // Събираме ВСИЧКИ кандидати и взимаме най-информативния.
    // Ако вземем първия заглавен елемент, за Кауфланд излиза само
    // марката („Martini“) вместо цялото име на продукта.
    const cands = [];
    el.querySelectorAll('[class*="title" i],[class*="name" i],[class*="titel" i],h2,h3,h4,h5')
      .forEach(t => cands.push((t.textContent || '').replace(/\s+/g, ' ').trim()));
    const img = el.querySelector('img');
    if (img && img.alt) {
      cands.push(img.alt.replace(ALT_PREFIX, '').replace(/\s+/g, ' ').trim());
    }
    (el.innerText || el.textContent || '').split('\n')
      .forEach(s => cands.push(s.replace(/\s+/g, ' ').trim()));

    let best = '';
    for (const s of cands) {
      if (!s || s.length < 3 || s.length > 140) continue;
      if (/^[\d.,\s%€лвkg\-–—]+$/i.test(s)) continue;   // само числа
      if (isUnitLine(s)) continue;
      if (looksLikeDescription(s)) continue;
      if (looksLikeCode(s)) continue;
      if (mostlyNumeric(s)) continue;
      if (s.length > best.length) best = s;
    }
    return best;
  };

  const imgOf = (el) => {
    const img = el.querySelector('img');
    if (img) {
      const d = img.currentSrc || img.src || img.getAttribute('data-src')
             || img.getAttribute('data-original') || '';
      if (d && d.indexOf('data:') !== 0) return {src: d, alt: (img.alt || '').trim()};
      const ss = img.getAttribute('data-srcset') || img.getAttribute('srcset') || '';
      if (ss) return {src: ss.split(',')[0].trim().split(' ')[0], alt: (img.alt || '').trim()};
      if (d) return {src: d, alt: (img.alt || '').trim()};
    }
    const s = el.querySelector('source');
    if (s) {
      const ss = s.getAttribute('srcset') || '';
      if (ss) return {src: ss.split(',')[0].trim().split(' ')[0], alt: ''};
    }
    return null;
  };

  const linkOf = (el) => {
    const a = el.tagName === 'A' ? el : (el.querySelector('a') || el.closest('a'));
    return a ? (a.href || '') : '';
  };

  const pctOf = (txt) => {
    const pm = txt.match(/-\s?(\d{1,2})\s?%/);
    return pm ? parseInt(pm[1], 10) : null;
  };

  const buildItem = (el) => {
    const t = (el.innerText || el.textContent || '').trim();
    const ms = pricesOf(el, t);
    if (!ms.length) return null;
    // Лидл слага цената в самото име: „Смядовска луканка XXL за 5.36 EUR“.
    // Махаме опашката, за да остане чисто име на продукта.
    let name = (nameOf(el) || '').replace(/\s+/g, ' ').trim();
    name = name.replace(/\s*[-–—]?\s*за\s+\d{1,4}[.,]\d{2}\s*(EUR|€|лв|BGN)\s*\.?$/i, '')
               .replace(/\s*\|\s*$/, '')
               .trim();
    if (!name || name.length < 3) return null;
    const im = imgOf(el);
    return {name: name,
            prices: ms.map(x => x.v),
            cur: ms.map(x => x.c),
            pct: pctOf(t),
            img: im ? im.src : '',
            url: linkOf(el)};
  };
"""

# --- 1) Самонастройващо се откриване на продуктовата решетка ------------
GRID_JS = "() => {" + JS_HELPERS + r"""
  // Продуктите в един списък са ЕДНАКВИ елементи, повторени много пъти.
  // Търсим class-подписа с най-много повторения, чиито елементи съдържат
  // цена — това е решетката, каквато и да е разметката.
  const sig = (el) => {
    const cls = (typeof el.className === 'string' ? el.className : '').trim();
    const parts = cls ? cls.split(/\s+/).filter(c => c.length > 1 &&
                        !/\d{3,}/.test(c)).slice(0, 3).sort().join('.') : '';
    return el.tagName + (parts ? '.' + parts : '');
  };

  const cands = [];
  const all = document.querySelectorAll('div,li,article,section,a');
  for (let i = 0; i < all.length; i++) {
    const el = all[i];
    if (el.children.length > 12) continue;
    const t = el.textContent || '';
    if (t.length < 4 || t.length > 400) continue;
    if (!PRICE_T.test(t)) continue;
    cands.push(el);
  }

  const groups = new Map();
  for (const el of cands) {
    let cur = el;
    for (let d = 0; d < 5 && cur && cur !== document.body; d++) {
      const s = sig(cur);
      if (s.length > 3) {
        if (!groups.has(s)) groups.set(s, new Set());
        groups.get(s).add(cur);
      }
      cur = cur.parentElement;
    }
  }

  let best = null, bestScore = 0, bestSig = '';
  for (const [s, set] of groups) {
    const els = Array.from(set);
    if (els.length < 6) continue;
    let sum = 0, imgs = 0;
    for (const e of els) {
      sum += (e.textContent || '').length;
      if (e.querySelector('img,source')) imgs++;
    }
    const avg = sum / els.length;
    if (avg > 400) continue;
    const score = els.length * (avg < 220 ? 2 : 1) * (imgs > els.length / 2 ? 2 : 1);
    if (score > bestScore) { bestScore = score; best = els; bestSig = s; }
  }
  if (!best) return {items: [], sig: '', n: 0};

  const out = [], seen = new Set();
  for (const el of best) {
    const it = buildItem(el);
    if (!it) continue;
    const key = it.name.toLowerCase().slice(0, 60) + '|' + Math.min.apply(null, it.prices);
    if (seen.has(key)) continue;
    seen.add(key);
    out.push(it);
  }
  return {items: out, sig: bestSig, n: best.length};
}
"""

# --- 2) Резервен метод: по плочки --------------------------------------
TILE_JS = "() => {" + JS_HELPERS + r"""
  const hasImage = (el) => !!(el.querySelector('img') || el.querySelector('source'));
  const cands = new Set();
  document.querySelectorAll('a').forEach(a => { if (hasImage(a)) cands.add(a); });
  document.querySelectorAll(
    '[class*="product"],[class*="Product"],[class*="tile"],[class*="Tile"],' +
    '[class*="offer"],[class*="Offer"],[class*="article"],[data-testid*="product"],' +
    'article,li'
  ).forEach(el => {
    const t = el.innerText || '';
    if (hasImage(el) && t.length > 0 && t.length < 400) cands.add(el);
  });

  const out = [], seen = new Set();
  cands.forEach(el => {
    const it = buildItem(el);
    if (!it) return;
    if (it.prices.length > 5) return;      // групов контейнер, не плочка
    const key = it.name.toLowerCase().slice(0, 60) + '|' + Math.min.apply(null, it.prices);
    if (seen.has(key)) return;
    seen.add(key);
    out.push(it);
  });
  return out;
}
"""

# За всяка верига: начални страници + думи, по които се разпознават
# връзките към още страници с оферти/категории. Обхождат се до MAX_PAGES.
#
# Списъкът се чете от config.json (ключ "stores"), за да се добавят вериги
# БЕЗ промяна в кода. Долното е само резервният вариант, ако го няма там.
DEFAULT_TARGETS = [
    ("Кауфланд",
     ["https://www.kaufland.bg/aktualni-predlozheniya/ot-ponedelnik.html",
      "https://www.kaufland.bg/aktualni-predlozheniya/oferti.html"],
     ("aktualni-predlozheniya", "oferti", "ot-ponedelnik", "ot-chetvartak")),
    ("Лидл",
     ["https://www.lidl.bg/"],
     ("aktsiya", "oferti", "predlozheni", "top-", "/c/")),
    ("Метро",
     ["https://www.metro.bg/oferti/top-oferti", "https://www.metro.bg/"],
     ("oferti", "promo", "katalog")),
    ("Билла",
     ["https://www.billa.bg/promocii"],
     ("promocii", "promo", "oferti", "katalog")),
]


# Колко страници да се обходят за конкретна верига. Има смисъл, защото
# Кауфланд дава 1108 оферти от 3 страници, а Лидл обикаля 20 и връща 64:
# без таван бавната верига изяжда бюджета на бързите.
STORE_PAGES = {}


def _load_targets():
    """Чете веригите от config.json; при липса ползва вградените."""
    raw = CFG.get("stores")
    if not raw:
        return DEFAULT_TARGETS
    out = []
    for s in raw:
        try:
            if not s.get("enabled", True):
                continue
            name = str(s["name"]).strip()
            urls = [u for u in s.get("urls", []) if str(u).startswith("http")]
            kws = tuple(str(k).lower() for k in s.get("keywords", []))
            if s.get("max_pages"):
                STORE_PAGES[name] = int(s["max_pages"])
            if name and urls:
                out.append((name, urls, kws))
        except Exception as e:
            log_err("config/stores", e)
    return out or DEFAULT_TARGETS


TARGETS = _load_targets()

# Лимитите се четат от config.json, за да се настройват без промяна на кода.
import time as _time
_START = _time.time()
# Общ времеви бюджет за обхождането. GitHub Actions спира задачата по
# някое време; ако се ударим в лимита, оставаме БЕЗ никакъв изход. Затова
# спираме сами навреме и записваме каквото сме събрали дотук.
TIME_BUDGET_S = int(CFG.get("time_budget_min", 45)) * 60


def out_of_time():
    return (_time.time() - _START) > TIME_BUDGET_S


def elapsed_min():
    return (_time.time() - _START) / 60.0


MAX_PAGES = int(CFG.get("max_pages", 60))          # страници на верига
MAX_PER_STORE = int(CFG.get("max_per_store", 1500))  # оферти на верига в изхода
LOAD_MORE_ROUNDS = int(CFG.get("load_more_rounds", 25))  # натискания на „покажи още“
POLITE_MS = int(CFG.get("polite_ms", 400))         # пауза между страниците

# Бутони за дозареждане — по текст, защото класовете се сменят постоянно.
LOAD_MORE_JS = """
() => {
  const words = ['покажи още','зареди още','още продукти','виж още','load more',
                 'показать','повече','следваща страница'];
  const els = [...document.querySelectorAll('button,a,div[role=button],span[role=button]')];
  for (const el of els) {
    const t = (el.innerText || el.textContent || '').trim().toLowerCase();
    if (!t || t.length > 40) continue;
    if (words.some(w => t.includes(w))) {
      const r = el.getBoundingClientRect();
      if (r.width > 0 && r.height > 0) { el.scrollIntoView({block:'center'}); el.click(); return true; }
    }
  }
  return false;
}
"""

# Колко продуктови елемента има в момента — за да знаем дали расте.
COUNT_JS = """
() => {
  const P = /\\d{1,4}[.,]\\d{2}/;
  let n = 0;
  const all = document.querySelectorAll('div,li,article,a');
  for (const el of all) {
    if (el.children.length > 12) continue;
    const t = el.textContent || '';
    if (t.length > 3 && t.length < 400 && P.test(t)) n++;
  }
  return n;
}
"""

# Връзки към следващи страници (pagination)
NEXT_JS = """
() => {
  const out = new Set();
  const rel = document.querySelector('link[rel=next], a[rel=next]');
  if (rel && rel.href) out.add(rel.href);
  document.querySelectorAll('a[href]').forEach(a => {
    const h = a.href || '';
    const t = (a.innerText || '').trim().toLowerCase();
    if (/[?&](page|p|strana|offset)=\\d+/i.test(h)) out.add(h.split('#')[0]);
    else if (t === 'следваща' || t === 'напред' || t === 'next' || t === '>') {
      if (h) out.add(h.split('#')[0]);
    }
  });
  return [...out];
}
"""


# Страници, които никога не съдържат продукти. Обхождането им само яде
# време и пълни „errors“ — при Лидл 15 от 20 посетени страници бяха такива.
# Дълги и характерни — търсят се като подниз
SKIP_URL_PART = ("zashchita-na-dannite", "obshti-uslovia", "polezno-info",
                 "magazin-lokator", "newsletter", "impressum", "privacy-policy",
                 "cookie-policy", "chesto-zadavani", "dostavka-i-plashtane",
                 # dm вика „тествайте продукт“ страници продуктови, но там
                 # няма цени — числата по тях са дати и брой участници.
                 # Оттам идваха „Печеливши участници −15%“ и „Тествайте…“.
                 "produkt-test", "test-na-produkt", "produktov-svyat",
                 "nashata-otgovornost", "za-dm-kontserna", "informatsiya-za-",
                 # редакционни/корпоративни рубрики без стоки
                 "novini", "blog", "recepti", "sertifikat")
# Кратки — само като ЦЯЛА част от адреса. Иначе „app“ съвпада вътре в
# случайни думи и изяжда цели вериги: точно така dm падна от 222 на 0.
SKIP_URL_WORD = {"uslovia", "privacy", "cookie", "cookies", "karier", "kariera",
                 "rabota", "kontakt", "kontakti", "za-nas", "igra", "igri",
                 "kupon", "kuponi", "coupid", "plus", "app", "lokator",
                 "terms", "about", "careers", "contact"}

SKIP_TITLE = ("защита на данните", "общи условия", "поверителност", "бисквитк",
              "кариер", "контакт", "за нас", "полезно инфо", "доставка и плащане",
              "често задавани", "карта на сайта")


def _useful_url(url):
    low = url.lower()
    if any(k in low for k in SKIP_URL_PART):
        return False
    parts = set(re.split(r"[/\-_.?=&#]+", low))
    return not (parts & SKIP_URL_WORD)


def _harvest(ctx, store, url, keywords=()):
    pg = ctx.new_page()
    found, extra_links = [], []
    try:
        pg.goto(url, wait_until="domcontentloaded", timeout=60000)
        pg.wait_for_timeout(3000)

        # Защитен екран („Един момент…“, „Checking your browser“) — сайтът
        # още не е зареден. Изчакваме проверката да мине, вместо да четем
        # междинната страница и да отчетем 0 продукта.
        # 5 опита по 2.5 сек стигат за нормална проверка; ако сайтът не
        # пусне дотогава, той просто блокира автоматизацията и чакането
        # само изяжда бюджета (миналото пускане отиде 28 мин заради това).
        for _ in range(5):
            try:
                txt = (pg.inner_text("body") or "")
                title = (pg.title() or "")
            except Exception:
                break
            waiting = (len(txt) < 600 or
                       re.search(r"един момент|checking your browser|just a moment|"
                                 r"проверка на браузъра|attention required",
                                 txt + " " + title, re.IGNORECASE))
            if not waiting:
                break
            pg.wait_for_timeout(2500)

        pg.wait_for_timeout(1500)
        # --- Дозареждане: скролваме и натискаме „покажи още“, докато расте ---
        last, stale = -1, 0
        for _ in range(LOAD_MORE_ROUNDS):
            for _ in range(3):
                pg.mouse.wheel(0, 3000)
                pg.wait_for_timeout(500)
            clicked = False
            try:
                clicked = bool(pg.evaluate(LOAD_MORE_JS))
            except Exception:
                pass
            pg.wait_for_timeout(1200 if clicked else 600)
            try:
                now = int(pg.evaluate(COUNT_JS))
            except Exception:
                break
            if now <= last:
                stale += 1
                if stale >= 2 and not clicked:
                    break          # два пъти подред нищо ново — стигнали сме дъното
            else:
                stale = 0
            last = now

        # Страница с условия/кариери няма продукти — не си губим времето
        try:
            ttl = (pg.title() or "").lower()
            if any(k in ttl for k in SKIP_TITLE):
                return [], []
        except Exception:
            pass

        # 1) самонастройващо се откриване на продуктовата решетка
        items, used = [], ""
        try:
            grid = pg.evaluate(GRID_JS) or {}
            items = grid.get("items") or []
            if items:
                used = f"решетка {grid.get('sig','?')} ({grid.get('n',0)} елемента)"
        except Exception as e:
            log_err(f"grid/{store}", e)

        # 2) резервно: старият метод по плочки
        if len(items) < 5:
            try:
                alt = pg.evaluate(TILE_JS) or []
                if len(alt) > len(items):
                    items, used = alt, "плочки (резервен метод)"
            except Exception as e:
                log_err(f"tiles/{store}", e)

        if used:
            OUT["stats"][f"метод/{store}"] = used

        # Връзки към още страници — категории + пагинация. Така стигаме до
        # целия каталог, а не само до първата страница.
        host = re.sub(r"^https?://([^/]+).*$", r"\1", url)
        seen = set()
        try:                                    # следваща страница (pagination)
            for l in (pg.evaluate(NEXT_JS) or []):
                if host in l and l not in seen:
                    seen.add(l)
                    extra_links.append(l)
        except Exception:
            pass
        if keywords:                            # категории и раздели с оферти
            try:
                links = pg.evaluate(
                    "() => [...document.querySelectorAll('a[href]')].map(a => a.href)")
                for l in links:
                    if host not in l or l in seen or not _useful_url(l):
                        continue
                    if any(k in l.lower() for k in keywords):
                        seen.add(l)
                        extra_links.append(l.split("#")[0])
            except Exception:
                pass

        raw = len(items)
        if raw == 0:
            # Диагностика на място: без нея „0 оферти“ не казва нищо.
            try:
                diag = pg.evaluate(
                    "() => ({title: document.title,"
                    " links: document.querySelectorAll('a').length,"
                    " imgs: document.querySelectorAll('img').length,"
                    " text: (document.body.innerText||'').length})")
                note(f"dom/{store}", f"0 плочки на {url[:60]} — "
                                     f"страница „{str(diag.get('title'))[:40]}“, "
                                     f"{diag.get('links')} линка, {diag.get('imgs')} картинки, "
                                     f"{diag.get('text')} знака текст")
            except Exception:
                note(f"dom/{store}", f"0 плочки на {url[:60]} (и диагностиката не мина)")

        for it in items:
            # първо махаме дублите в другата валута, после избираме двойката
            prices = collapse_currency(it.get("prices", []),
                                       it.get("cur"), it.get("pct"))
            prices = [x for x in prices if 0.1 <= x <= 5000]
            if not prices:
                continue
            name = re.sub(r"\s+", " ", it.get("name") or "").strip()
            if sum(c.isalpha() for c in name) < 5:
                continue
            price, old = pick_price_pair(prices, it.get("pct"))
            if price is None:
                continue
            rec = {
                "store": store, "name": name[:90], "price": round(price, 2),
                "old": round(old, 2) if old else None,
                "img": (it.get("img") or "")[:300],
                "url": (it.get("url") or "")[:300],
            }
            # Процентът, изписан на самата плочка, е по-верен от сметнатия:
            # понякога на плочката има и цена за килограм, която обърква сметката.
            if it.get("pct"):
                rec["pct"] = int(it["pct"])
            found.append(enrich(rec))
        if raw and not found:
            note(f"dom/{store}", f"{raw} плочки, но 0 с валидна цена/име")
        return found, extra_links
    except Exception as e:
        log_err(f"dom/{store}/{url[:45]}", e)
        return found, []
    finally:
        try:
            pg.close()
        except Exception:
            pass


def scrape_dom_offers(ctx):
    if ctx is None:
        note("dom", "браузърът не е наличен — офертите се пропускат")
        return
    for store, urls, keywords in TARGETS:
        if out_of_time():
            note("време", f"бюджетът свърши — {store} и следващите са пропуснати")
            break
        # Времето се дели поравно между веригите, за да не изяде първата всичко
        store_deadline = _time.time() + max(60, (TIME_BUDGET_S - (_time.time() - _START))
                                            / max(1, len(TARGETS) - TARGETS.index((store, urls, keywords))))
        page_cap = STORE_PAGES.get(store, MAX_PAGES)
        per_store, queue, visited = [], list(urls), set()
        while queue and len(visited) < page_cap:
            if out_of_time() or _time.time() > store_deadline:
                OUT["stats"][f"време/{store}"] = f"спрян на {len(visited)} страници"
                break
            url = queue.pop(0)
            if url in visited:
                continue
            visited.add(url)
            items, extra = _harvest(ctx, store, url, keywords)
            per_store += items
            for l in extra:
                if l not in visited and l not in queue and len(queue) < page_cap * 3:
                    queue.append(l)
            # Малка пауза — не искаме да натоварваме чужди сайтове
            if POLITE_MS:
                try:
                    ctx.pages[0].wait_for_timeout(POLITE_MS)
                except Exception:
                    pass
            # Достатъчно събрахме за тази верига
            if len(per_store) >= MAX_PER_STORE * 2:
                break

        seen, ded, junk = set(), [], 0
        for o in per_store:
            if is_junk_offer(o):
                junk += 1
                continue
            o = clean_offer(o)
            k = (o["name"].lower(), o["price"])
            if k not in seen:
                seen.add(k)
                ded.append(o)
        OUT["offers"] += ded[:MAX_PER_STORE]
        OUT["stats"][f"offers/{store}"] = f"{len(ded[:MAX_PER_STORE])} от {len(visited)} страници"
        if junk:
            OUT["stats"][f"изхвърлени/{store}"] = f"{junk} нестокови записа"
        if not ded:
            note(f"dom/{store}", "0 оферти — вероятно смениха сайта, провери селекторите")

    OUT["brochures"] = [o for o in OUT["offers"] if o["store"] in ("Лидл", "Билла")]


# ---------------------------------------------------------------------------
# 3. kolkostruva.bg — официалните дневни цени
# ---------------------------------------------------------------------------
CHAIN_MAP = [
    ("КАУФЛАНД", "Кауфланд"), ("KAUFLAND", "Кауфланд"),
    ("БИЛЛА", "Билла"), ("BILLA", "Билла"),
    ("ЛИДЛ", "Лидл"), ("LIDL", "Лидл"),
    ("МЕТРО", "Метро"), ("METRO", "Метро"),
    ("ФАНТАСТИКО", "Фантастико"), ("FANTASTIKO", "Фантастико"),
    ("Т МАРКЕТ", "T-Market"), ("T MARKET", "T-Market"), ("ТМАРКЕТ", "T-Market"),
    ("БУЛМАГ", "BulMag"), ("BULMAG", "BulMag"),
    ("БЕРЕЗКА", "Березка"), ("BEREZKA", "Березка"),
    ("АВАНТИ", "Аванти"), ("AVANTI", "Аванти"),
    ("ЛЕКСИ", "Лекси"), ("LEKSI", "Лекси"),
    ("КЛАСИКО", "Класико"), ("KLASIKO", "Класико"),
    ("ПРОМАРКЕТ", "Promarket"), ("ТРИУМФ", "Триумф"),
    ("ЖАНЕТ", "Жанет"), ("ПИКАДИЛИ", "Пикадили"), ("PICCADILLY", "Пикадили"),
    ("ЛИЛИ", "Lilly"), ("LILLY", "Lilly"),
    ("ЕБАГ", "eBag"), ("EBAG", "eBag"),
    # Кратките ключове се проверяват с граници на думата (виж _key_hit),
    # за да не хванат „ДАР“ вътре в „БОЖУРДАР“.
    ("CBA", "CBA"), ("СБА", "CBA"), ("ДАР", "Дар"), ("DAR", "Дар"),
    ("DM", "dm"), ("ДМ", "dm"),
]

# Аптеките отпадат — там се продават лекарства, не хранителни стоки.
# Дрогериите (Lilly, dm) ОСТАВАТ: там има перилни, козметика и битова химия,
# които влизат в кошницата. Конкуренцията ги показва и с право.
# „АПТЕК“ без окончание — иначе „АПТЕКИ ГАЛЕН“ се промъква покрай „АПТЕКА“.
SKIP_CHAIN = ("АПТЕК", "APTEK", "ФАРМА", "PHARM", "ЛЕКАРСТВ", "ЗДРАВНА КАСА",
              "ОПТИК", "ЗООМАГ",
              # бензиностанции: продават храна, но цените им не са
              # представителни за пазаруване и само шумят в класацията
              "БЕНЗИНОСТ", "ПЕТРОЛ", "PETROL", "ЛУКОЙЛ", "LUKOIL",
              "ШЕЛ", "SHELL", "ОМВ", "OMV", "ЕКО БЪЛГАРИЯ", "ROMPETROL")


def _key_hit(up, key):
    """Дълъг ключ = обикновено съвпадение; кратък = само като цяла дума."""
    if len(key) >= 5:
        return key in up
    return re.search(r"(?<![А-ЯЁA-Z])" + re.escape(key) + r"(?![А-ЯЁA-Z])", up) is not None


def norm_chain(raw):
    up = raw.upper()
    if any(k in up for k in SKIP_CHAIN):
        return None
    for key, nice in CHAIN_MAP:
        if _key_hit(up, key):
            return nice
    # Непозната верига: чистим я, но я ПАЗИМ — държавните данни покриват
    # и вериги, които никой не е сложил в списък, и точно там е предимството.
    clean = re.sub(r"_\d+$", "", raw).strip()
    clean = clean.split("(")[0].strip()          # маха юридическите опашки
    clean = re.sub(r"\s+(ЕООД|ООД|АД|ЕАД|ЕТ|КД|СД)\b.*$", "", clean,
                   flags=re.IGNORECASE).strip()
    clean = re.sub(r"\s{2,}", " ", clean)
    return clean[:24] if len(clean) >= 2 else None


def _read_csv(raw, chain, rows):
    for enc in ("utf-8-sig", "windows-1251", "utf-8"):
        try:
            textdata = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        return
    first = textdata.splitlines()[0] if textdata else ""
    delim = max([";", ",", "\t"], key=first.count)
    rdr = csv.reader(io.StringIO(textdata), delimiter=delim)
    hdr = [h.strip().lower() for h in next(rdr, [])]

    def col(*names):
        for i, h in enumerate(hdr):
            if any(n in h for n in names):
                return i
        return -1

    pc = col("продукт", "стока", "наименование", "артикул", "product", "name")
    cc = col("цена", "price")
    if pc < 0 or cc < 0:
        note("basics", f"{chain[:30]}: колони продукт/цена не се разпознават ({hdr[:4]})")
        return
    for row in rdr:
        try:
            price = float(re.sub(r"[^\d.]", "", row[cc].replace(",", ".")))
            nmv = row[pc].strip()
            ch = norm_chain(chain)
            if ch and nmv and 0 < price < 1000:
                rows.append(enrich({"chain": ch, "product": nmv,
                                    "name": nmv, "price": price}))
        except Exception:
            pass


def _read_xlsx(blob, chain, rows):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(blob), read_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    hdr = [str(h or "").strip().lower() for h in next(it, [])]

    def col(*names):
        for i, h in enumerate(hdr):
            if any(n in h for n in names):
                return i
        return -1

    pc = col("продукт", "стока", "наименование", "артикул")
    cc = col("цена")
    if pc < 0 or cc < 0:
        return
    for row in it:
        try:
            price = float(str(row[cc]).replace(",", "."))
            nmv = str(row[pc] or "").strip()
            ch = norm_chain(chain)
            if ch and nmv and 0 < price < 1000:
                rows.append(enrich({"chain": ch, "product": nmv,
                                    "name": nmv, "price": price}))
        except Exception:
            pass


# Веригите, в които хората реално пазаруват. Те получават дял ПЪРВИ —
# иначе се пълни с местни магазинчета и в приложението липсват Кауфланд
# и Билла, което прави класацията на кошницата безполезна.
PRIORITY_CHAINS = [
    "Кауфланд", "Билла", "Лидл", "Метро", "Фантастико", "T-Market",
    "BulMag", "dm", "Lilly", "CBA", "Аванти", "КООП", "Триумф",
    "Промаркет", "Пикадили", "Жанет", "Березка", "Дар", "Лекси", "Класико",
]


# Кошницата: с какво реално се пазарува. Официалният архив на КЗП е над
# милион реда и ако се реже сляпо, в приложението остават случайни стоки от
# случайни вериги — точно затова „яйца“ намираше само „ГРОС МАКАРОНИ С ЯЙЦА“
# в три непознати магазина. Тук всяка верига получава своя дял ПЪРВО от
# основните стоки, и чак после от останалото.
BASKET = {
    "хляб": ("ХЛЯБ", "ХЛЕБ", "ПИТК"),
    "мляко": ("МЛЯКО", "МЛЕК"),
    "кисело мляко": ("КИСЕЛО",),
    "сирене": ("СИРЕНЕ", "СИРЕН"),
    "кашкавал": ("КАШКАВАЛ",),
    "масло": ("МАСЛО",),
    "олио": ("ОЛИО",),
    "яйца": ("ЯЙЦА", "ЯЙЦЕ"),
    "захар": ("ЗАХАР",),
    "брашно": ("БРАШНО",),
    "ориз": ("ОРИЗ",),
    "макарони": ("МАКАРОН", "СПАГЕТ", "ФИДЕ", "ПАСТА"),
    "боб": ("БОБ", "ФАСУЛ"),
    "леща": ("ЛЕЩА",),
    "картофи": ("КАРТОФ",),
    "домати": ("ДОМАТ",),
    "краставици": ("КРАСТАВИЦ",),
    "лук": ("ЛУК",),
    "чушки": ("ЧУШК", "ПИПЕРК"),
    "моркови": ("МОРКОВ",),
    "зеле": ("ЗЕЛЕ",),
    "ябълки": ("ЯБЪЛК",),
    "банани": ("БАНАН",),
    "портокали": ("ПОРТОКАЛ",),
    "лимони": ("ЛИМОН",),
    "грозде": ("ГРОЗДЕ",),
    "диня": ("ДИНЯ", "ПЪПЕШ"),
    "кайма": ("КАЙМА",),
    "свинско": ("СВИНС", "СВИНСКО"),
    "пилешко": ("ПИЛЕ", "ПИЛЕШ"),
    "телешко": ("ТЕЛЕШ", "ГОВЕЖД"),
    "кебапче": ("КЕБАПЧ", "КЮФТЕ"),
    "наденица": ("НАДЕНИЦ",),
    "луканка": ("ЛУКАНК",),
    "салам": ("САЛАМ",),
    "шунка": ("ШУНКА",),
    "риба": ("РИБА", "СКУМРИЯ", "СЬОМГА", "ХЕК", "ЦАЦА", "ПЪСТЪРВ"),
    "кафе": ("КАФЕ",),
    "чай": ("ЧАЙ",),
    "вода": ("ВОДА",),
    "сок": ("СОК", "НЕКТАР"),
    "бира": ("БИРА",),
    "вино": ("ВИНО",),
    "ракия": ("РАКИЯ", "ВОДКА", "УИСКИ", "МАСТИКА"),
    "оцет": ("ОЦЕТ",),
    "сол": ("СОЛ",),
    "черен пипер": ("ПОДПРАВК", "ЧУБРИЦ"),
    "шоколад": ("ШОКОЛАД",),
    "вафла": ("ВАФЛ",),
    "бисквити": ("БИСКВИТ",),
    "чипс": ("ЧИПС", "СНАКС", "КРЕКЕР"),
    "ядки": ("ЯДК", "ФЪСТЪ", "БАДЕМ", "КАШУ", "ЛЕШНИК"),
    "сладолед": ("СЛАДОЛЕД",),
    "мед": ("МЕД",),
    "конфитюр": ("КОНФИТЮР", "МАРМАЛАД"),
    "тахан": ("ТАХАН", "ХАЛВА"),
    "кетчуп": ("КЕТЧУП",),
    "майонеза": ("МАЙОНЕЗА",),
    "горчица": ("ГОРЧИЦА", "ЛЮТЕНИЦ"),
    "тоалетна хартия": ("ТОАЛЕТНА",),
    "салфетки": ("САЛФЕТК", "КЪРПИЧК"),
    "прах за пране": ("ПРАХ", "ПЕРИЛЕН", "ГЕЛ ЗА ПРАНЕ"),
    "омекотител": ("ОМЕКОТИТЕЛ",),
    "препарат за съдове": ("ПРЕПАРАТ", "ПОЧИСТВАЩ"),
    "сапун": ("САПУН",),
    "шампоан": ("ШАМПОАН",),
    "паста за зъби": ("ПАСТА ЗА ЗЪБИ", "ЗЪБНА"),
    "дезодорант": ("ДЕЗОДОРАНТ", "АНТИПЕРСПИРАНТ"),
    "пелени": ("ПЕЛЕН",),
}

# Думи, след които следващата дума е СЪСТАВКА, а не стоката. Без тях
# „яйца“ лови „макарони С яйца“, а „мляко“ лови „шоколад С мляко“.
_INGREDIENT_MARKER = {"С", "СЪС", "ВКУС", "АРОМАТ", "ЗА", "БЕЗ", "ОТ",
                      "И", "В", "ВЪВ", "НА", "ПЪЛНЕЖ", "ПЛЪНКА"}


# Грамажите не са част от името: „450 ГР. СЕЛСКИ ПШЕНИЧЕН ХЛЯБ“ трябва да
# се брои за хляб, а мярката отпред само измества думата надясно.
_UNIT_WORDS = {"Г", "ГР", "КГ", "МЛ", "Л", "ML", "G", "KG", "L",
               "БР", "БРОЯ", "Х", "X", "СМ", "CM"}


def _name_words(s):
    out = []
    for w in re.split(r"[^0-9A-Za-zА-Яа-яЁё]+", (s or "").upper()):
        if not w or w[0].isdigit() or w in _UNIT_WORDS:
            continue
        out.append(w)
    return out


def basket_key(name):
    """Към коя основна стока спада редът. None = не е от кошницата.

    Гледа се и КЪДЕ стои думата: стоката се казва в началото на името,
    а съставките — след „с“/„със“. Затова „ЯЙЦА ЗДРАВКОВЕЦ 6БР“ минава,
    а „ГРОС МАКАРОНИ С ЯЙЦА“ — не.
    """
    ws = _name_words(name)
    if not ws:
        return None
    best = None
    for key, stems in BASKET.items():
        for i, w in enumerate(ws[:4]):
            if not any(w.startswith(st) for st in stems):
                continue
            if i and ws[i - 1] in _INGREDIENT_MARKER:
                break
            if best is None or i < best[1]:
                best = (key, i)
            break
    return best[0] if best else None


def _select_basics(rows):
    """Избира кои официални цени да влязат във фийда.

    Архивът на КЗП е над милион реда от 128 вериги — по един запис за всеки
    магазин. Старият подход (първите N реда на верига) даваше файл, в който
    големите вериги ги нямаше, а обикновените стоки ги имаше само на случаен
    принцип. Затова сега:

      1) сливане — една цена за продукт във верига, най-ниската;
      2) кошницата ПЪРВА — за всяка верига и всяка основна стока се пазят
         най-евтините няколко реда, тоест всяка верига има мляко, яйца, хляб;
      3) остатъкът пълни каквото е останало от тавана, приоритетно за
         веригите, в които хората реално пазаруват.
    """
    per_chain = int(CFG.get("max_basics_per_chain", 1500))
    total_cap = int(CFG.get("max_basics_total", 26000))
    per_item = int(CFG.get("max_basics_per_item", 4))
    basket_cap = int(CFG.get("max_basics_basket", 22000))

    # 1) сливане: най-ниската цена за продукт във всяка верига
    merged = {}
    for r in rows:
        key = (r["chain"], re.sub(r"\s+", " ", r["product"]).strip().lower()[:70])
        cur = merged.get(key)
        if cur is None or r["price"] < cur["price"]:
            merged[key] = r

    # 2) разделяне на „кошница“ и „останало“
    pairs, rest = {}, {}
    for r in merged.values():
        k = basket_key(r["product"])
        if k:
            r["k"] = k
            pairs.setdefault((r["chain"], k), []).append(r)
        else:
            rest.setdefault(r["chain"], []).append(r)

    def rank(chain):
        return (PRIORITY_CHAINS.index(chain) if chain in PRIORITY_CHAINS
                else len(PRIORITY_CHAINS) + 1)

    picked = []
    for (chain, k) in sorted(pairs.keys(), key=lambda ck: (rank(ck[0]), ck[1])):
        if len(picked) >= basket_cap:
            break
        lst = sorted(pairs[(chain, k)], key=lambda r: r["price"])
        picked.extend(lst[:per_item])
    basket_n = len(picked)

    # 3) останалото — по приоритет на веригата, докато има място
    order = sorted(rest.keys(), key=lambda c: (rank(c), -len(rest[c])))
    room_per_chain = max(50, per_chain)
    for ch in order:
        if len(picked) >= total_cap:
            break
        picked.extend(sorted(rest[ch], key=lambda r: r["price"])[:room_per_chain])

    picked = picked[:total_cap]

    # „name“ дублира „product“ дума по дума — маха се, това е 30% от файла
    for r in picked:
        r.pop("name", None)

    got = sorted({r["chain"] for r in picked})
    basket_chains = sorted({r["chain"] for r in picked if r.get("k")})
    OUT["stats"]["basics_вериги_в_архива"] = len({r["chain"] for r in merged.values()})
    OUT["stats"]["basics_редове_в_архива"] = len(rows)
    OUT["stats"]["basics_след_сливане"] = len(merged)
    OUT["stats"]["basics"] = (f"{len(picked)} реда, {len(got)} вериги "
                              f"(от тях {basket_n} основни стоки в "
                              f"{len(basket_chains)} вериги)")
    OUT["stats"]["basics_chains"] = got
    missing = [c for c in PRIORITY_CHAINS[:6] if c not in got]
    if missing:
        note("basics", f"липсват едри вериги: {', '.join(missing)} — "
                       f"или не подават данни, или името им не се разпознава")
    return picked


def scrape_basics():
    base = CFG.get("food_base", "https://kolkostruva.bg/opendata_files/")
    day = datetime.date.today()
    for back in range(8):
        d = (day - datetime.timedelta(days=back)).isoformat()
        try:
            r = requests.get(f"{base}{d}.zip", headers=UA, timeout=60)
            if r.status_code != 200 or len(r.content) < 1000:
                continue
            zf = zipfile.ZipFile(io.BytesIO(r.content))
            rows, skipped = [], 0
            for nm in zf.namelist():
                low = nm.lower()
                chain = nm.rsplit("/", 1)[-1].rsplit(".", 1)[0]
                if norm_chain(chain) is None:
                    skipped += 1
                    continue                      # аптеки и пр. — не ги четем изобщо
                try:
                    if low.endswith(".csv"):
                        _read_csv(zf.read(nm), chain, rows)
                    elif low.endswith((".xlsx", ".xls")):
                        _read_xlsx(zf.read(nm), chain, rows)
                except Exception as e:
                    log_err(f"basics/{nm[:40]}", e)
            if rows:
                OUT["basics"] = _select_basics(rows)
                OUT["basics_date"] = d
                OUT["stats"]["basics_пропуснати_файла"] = skipped
                return
        except Exception as e:
            log_err(f"basics/{d}", e)
    note("basics", "нито един архив от последните 8 дни не се прочете")


# ---------------------------------------------------------------------------
# Разделяне на изхода: телефонът да тегли само каквото гледа
# ---------------------------------------------------------------------------
SLUG_MAP = {
    "Кауфланд": "kaufland", "Лидл": "lidl", "Билла": "billa", "Метро": "metro",
    "Фантастико": "fantastiko", "T-Market": "tmarket", "BulMag": "bulmag",
    "Березка": "berezka", "Аванти": "avanti", "Дар": "dar", "Лекси": "leksi",
    "Класико": "klasiko", "Lilly": "lilly", "dm": "dm", "CBA": "cba",
    "eBag": "ebag", "Промаркет": "promarket", "Триумф": "triumf",
    "Жанет": "janet", "Пикадили": "piccadilly",
}


def slug(name):
    if name in SLUG_MAP:
        return SLUG_MAP[name]
    s = re.sub(r"[^a-zA-Zа-яА-Я0-9]+", "-", name).strip("-").lower()
    tr = str.maketrans("абвгдежзийклмнопрстуфхцчшщъьюя",
                       "abvgdejziyklmnoprstufhc4w6a-ya")
    return (s.translate(tr) or "store")[:20]


def _dump(path, obj):
    import os
    d = os.path.dirname(path)
    if d:
        os.makedirs(d, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, separators=(",", ":"))
    return os.path.getsize(path)


# ---------------------------------------------------------------------------
# 4. Цени, подадени от хора и от магазини
# ---------------------------------------------------------------------------
#
# Приложението няма сървър, затова подаването минава през чужда, безплатна
# кутия: Google формуляр пише отговорите в таблица, таблицата се публикува
# като CSV, а тук се чете само това, което е ОДОБРЕНО. Одобрението не е
# бюрокрация — без него конкурентът отсреща вписва лъжливи цени още първия
# ден и цялата база става безполезна.
#
# Ако адресът е празен, стъпката просто се пропуска.

COMMUNITY_COLS = {
    "shop": ("магазин", "shop", "обект"),
    "town": ("град", "town", "населено"),
    "product": ("продукт", "product", "стока", "артикул"),
    "price": ("цена", "price"),
    "qty": ("количество", "qty", "разфасовка", "грамаж"),
    "barcode": ("баркод", "barcode", "ean"),
    "date": ("дата", "date", "timestamp", "клеймо"),
    "approved": ("одобрен", "approved", "ок", "ok"),
    "role": ("роля", "role"),
}


def _col_index(header):
    """Кой стълб какво е — по име, а не по номер.

    Номерата се разместват при всяко добавяне на въпрос във формуляра;
    имената оцеляват.
    """
    idx = {}
    for i, h in enumerate(header):
        low = re.sub(r"\s+", " ", (h or "")).strip().lower()
        for key, names in COMMUNITY_COLS.items():
            if key in idx:
                continue
            if any(low.startswith(n) for n in names):
                idx[key] = i
    return idx


def _truthy(v):
    return str(v).strip().lower() in ("да", "yes", "true", "1", "x", "ок", "ok", "✓")


def scrape_community():
    url = (CFG.get("community_csv_url") or "").strip()
    if not url:
        return
    max_age = int(CFG.get("community_max_age_days", 21))
    require_ok = bool(CFG.get("community_require_approved", True))
    try:
        r = requests.get(url, headers=UA, timeout=45)
        if r.status_code != 200 or not r.content:
            note("общност", f"таблицата не се чете (HTTP {r.status_code})")
            return
        r.encoding = r.apparent_encoding or "utf-8"
        rows = list(csv.reader(io.StringIO(r.text)))
    except Exception as e:
        log_err("community", e)
        return

    if len(rows) < 2:
        note("общност", "таблицата е празна")
        return

    idx = _col_index(rows[0])
    missing = [k for k in ("shop", "product", "price") if k not in idx]
    if missing:
        note("общност", f"липсват стълбове: {', '.join(missing)} — "
                        f"провери имената на въпросите във формуляра")
        return

    today = datetime.date.today()
    out, skipped = [], {"неодобрени": 0, "стари": 0, "невалидни": 0}
    for row in rows[1:]:
        def cell(key):
            i = idx.get(key)
            return (row[i].strip() if i is not None and i < len(row) else "")

        if require_ok and not _truthy(cell("approved")):
            skipped["неодобрени"] += 1
            continue

        # Тук цената е гола („3.49“), не текст от плочка — PRICE_RE иска „лв“
        m = re.search(r"\d+(?:[.,]\d+)?", cell("price"))
        price = float(m.group(0).replace(",", ".")) if m else None
        name = re.sub(r"\s+", " ", cell("product")).strip()
        shop = re.sub(r"\s+", " ", cell("shop")).strip()
        if not name or not shop or price is None or price <= 0 or price > 5000:
            skipped["невалидни"] += 1
            continue

        # Свежест: стара подадена цена е по-лоша от липсваща, защото лъже
        d = cell("date")[:10]
        try:
            when = datetime.date.fromisoformat(d)
            if (today - when).days > max_age:
                skipped["стари"] += 1
                continue
        except Exception:
            when = None

        rec = {"chain": shop[:40], "product": name[:90], "price": round(price, 2),
               "town": cell("town")[:30], "date": d,
               "barcode": re.sub(r"\D", "", cell("barcode"))[:14],
               "src": "shop" if "магазин" in cell("role").lower() else "user"}
        qty = cell("qty")
        u = unit_price(rec["price"], f"{name} {qty}".strip())
        if u:
            rec["unitPrice"], rec["unitLabel"] = u
        out.append(rec)

    # Един магазин, един продукт — най-новата цена
    best = {}
    for rec in out:
        k = (rec["chain"].lower(), rec["product"].lower())
        cur = best.get(k)
        if cur is None or (rec.get("date") or "") >= (cur.get("date") or ""):
            best[k] = rec
    OUT["community"] = list(best.values())
    OUT["stats"]["общност"] = (f"{len(OUT['community'])} цени от "
                               f"{len({r['chain'] for r in OUT['community']})} магазина")
    for k, v in skipped.items():
        if v:
            OUT["stats"][f"общност_пропуснати_{k}"] = v


# ---------------------------------------------------------------------------
# 5. Доклади „цената не е вярна“
# ---------------------------------------------------------------------------
#
# Обхождането ще греши — това е неизбежно, когато данните идват от чужди
# сайтове, които се променят без предупреждение. Въпросът е дали грешката
# остава да лъже, или някой я маха.
#
# Тук докладите от хората стават предпазител: продукт, за който няколко
# НЕЗАВИСИМИ човека са казали едно и също, спира да се публикува. Прагът
# не е формалност — един доклад може да е грешка или злоба, три вече са
# данни. Записът не се трие мълчаливо: в stats пише колко и кои са паднали.

REPORT_COLS = {
    "shop": ("магазин", "shop", "верига"),
    "product": ("продукт", "product", "стока"),
    "kind": ("какво", "kind", "грешка", "вид"),
    "date": ("дата", "date", "timestamp", "клеймо"),
    "approved": ("одобрен", "approved", "ок", "ok"),
}


def _norm_key(s):
    return re.sub(r"[^0-9a-zA-Zа-яА-Я]+", " ", (s or "").lower()).strip()


def scrape_reports():
    """Сваля докладите и маха офертите, за които има достатъчно оплаквания."""
    url = (CFG.get("reports_csv_url") or "").strip()
    if not url:
        return
    min_count = int(CFG.get("reports_min_count", 3))
    max_age = int(CFG.get("reports_max_age_days", 30))
    try:
        r = requests.get(url, headers=UA, timeout=45)
        if r.status_code != 200 or not r.content:
            note("доклади", f"таблицата не се чете (HTTP {r.status_code})")
            return
        r.encoding = r.apparent_encoding or "utf-8"
        rows = list(csv.reader(io.StringIO(r.text)))
    except Exception as e:
        log_err("reports", e)
        return

    if len(rows) < 2:
        return

    hdr = rows[0]
    idx = {}
    for i, h in enumerate(hdr):
        low = re.sub(r"\s+", " ", (h or "")).strip().lower()
        for key, names in REPORT_COLS.items():
            if key not in idx and any(low.startswith(n) for n in names):
                idx[key] = i
    if "shop" not in idx or "product" not in idx:
        note("доклади", "липсват стълбове магазин/продукт")
        return

    today = datetime.date.today()
    counts = {}
    for row in rows[1:]:
        def cell(key):
            i = idx.get(key)
            return (row[i].strip() if i is not None and i < len(row) else "")

        # Неодобрените се броят също — докладът не е публикация, а сигнал.
        # Одобрението тук служи само да се СПРЕ явна злоупотреба.
        if idx.get("approved") is not None and \
                str(cell("approved")).strip().lower() in ("не", "no", "false", "0"):
            continue
        d = cell("date")[:10]
        try:
            if (today - datetime.date.fromisoformat(d)).days > max_age:
                continue
        except Exception:
            pass
        k = (_norm_key(cell("shop")), _norm_key(cell("product")))
        if not k[0] or not k[1]:
            continue
        counts[k] = counts.get(k, 0) + 1

    bad = {k for k, v in counts.items() if v >= min_count}
    if not bad:
        OUT["stats"]["доклади"] = (f"{len(counts)} докладвани продукта, "
                                   f"нито един не стига прага от {min_count}")
        return

    before = len(OUT["offers"])
    dropped = []
    kept = []
    for o in OUT["offers"]:
        k = (_norm_key(o["store"]), _norm_key(o["name"]))
        if k in bad:
            dropped.append(f"{o['store']}: {o['name'][:40]}")
        else:
            kept.append(o)
    OUT["offers"] = kept
    OUT["stats"]["доклади"] = (f"{len(counts)} докладвани продукта, "
                               f"{before - len(kept)} свалени (праг {min_count})")
    if dropped:
        OUT["stats"]["доклади_свалени"] = dropped[:20]


# ---------------------------------------------------------------------------
# 6. История на цените
# ---------------------------------------------------------------------------
#
# Обхождането вижда само днешния ден; историята се РЪСТИ от пускане на
# пускане: предишният feed/history.json се тегли от публикувания клон,
# днешните цени се добавят и файлът се записва обратно. Така графиката в
# приложението показва как се е движила цената — а това е и отговорът на
# най-важния въпрос: истинско ли е „намалението“, или цената първо беше
# вдигната.
#
# Точка се добавя само при ПРОМЯНА на цената (плюс първия запис) — иначе
# файлът расте с 1400 реда дневно, без да казва нищо ново. Продукти,
# невиждани от HISTORY_KEEP_DAYS дни, изпадат; точките се режат до
# HISTORY_MAX_POINTS на продукт.

HISTORY_MAX_POINTS = 40
HISTORY_KEEP_DAYS = 60


def _hist_key(store, name):
    return (store.strip().lower() + "|" +
            re.sub(r"\s+", " ", name).strip().lower()[:60])


def build_history():
    base = CFG.get(
        "feed_raw_base",
        "https://raw.githubusercontent.com/valerimilanov1990-max/promoradar-data"
    ).rstrip("/")

    prev = {}
    for br in ("data", "main"):
        try:
            r = requests.get(f"{base}/{br}/feed/history.json",
                             headers=UA, timeout=30)
            if r.status_code == 200 and r.content:
                prev = json.loads(r.text).get("h", {})
                break
        except Exception:
            continue

    today = datetime.date.today().isoformat()
    h = dict(prev)
    changed = 0
    for o in OUT["offers"]:
        k = _hist_key(o["store"], o["name"])
        e = h.get(k) or {"d": [], "p": [], "l": today}
        price = round(float(o["price"]), 2)
        if not e["p"] or abs(e["p"][-1] - price) > 0.005:
            e["d"].append(today)
            e["p"].append(price)
            changed += 1
        e["l"] = today
        e["d"] = e["d"][-HISTORY_MAX_POINTS:]
        e["p"] = e["p"][-HISTORY_MAX_POINTS:]
        h[k] = e

    # чистене: продукти, които никой не е виждал отдавна
    cutoff = (datetime.date.today()
              - datetime.timedelta(days=HISTORY_KEEP_DAYS)).isoformat()
    h = {k: v for k, v in h.items() if v.get("l", "") >= cutoff}

    OUT["history"] = h
    OUT["stats"]["история"] = (f"{len(h)} продукта, {changed} нови точки, "
                               f"наследени {len(prev)}")


# ---------------------------------------------------------------------------
# Локациите на магазините — за „вериги наблизо“ в приложението
# ---------------------------------------------------------------------------
#
# Източникът е OpenStreetMap (Overpass API): безплатен, без ключ, с добро
# покритие на веригите в България. Големите марки се разпознават по
# псевдоними (Kaufland → Кауфланд), кварталните — по точно съвпадение на
# името с верига от официалните данни. При недостъпен Overpass се наследява
# предишният файл — както при историята, локациите не се губят от един
# неуспешен опит. Координатите са с 4 знака (~11 м) — файлът остава малък.

GEO_ALIASES = {
    "Кауфланд": ["kaufland"],
    "Лидл": ["lidl"],
    "Билла": ["billa"],
    "Метро": ["metro", "метро"],
    "Фантастико": ["fantastico", "фантастико"],
    "T-Market": ["t market", "t-market", "т маркет", "т-маркет", "tmarket"],
    "dm": ["dm", "dm drogerie markt", "dm-drogerie markt"],
    "Lilly": ["lilly", "lilly drogerie", "лили", "лили дрогерие"],
    "BulMag": ["bulmag", "булмаг"],
    "КООП": ["кооп", "coop"],
    "Аванти": ["аванти", "avanti"],
    "ТРИСТА": ["триста", "магазини 345", "345"],
    "Жанет": ["жанет"],
    "Болеро": ["болеро"],
    "Макс": ["макс маркет"],
}

def _geo_norm(s):
    s = re.sub(r"[\"'„“”]+", "", s.lower())
    s = re.sub(r"[\s\-–—]+", " ", s).strip()
    return re.sub(r"^(супермаркет|магазин|магазини|market|супермаркети)\s+", "", s)

def _geo_prev():
    base = CFG.get("feed_raw_base", "").rstrip("/")
    if not base:
        return {}
    for br in ("data", "main"):
        try:
            r = requests.get(f"{base}/{br}/feed/stores-geo.json", timeout=30)
            if r.ok:
                return r.json().get("chains", {})
        except Exception:
            pass
    return {}

def build_geo():
    chains = ({o["store"] for o in OUT["offers"]}
              | {b["chain"] for b in OUT["basics"]}
              | {"Билла", "Метро", "Фантастико", "T-Market", "Lilly"})
    # нормализирано име → име на верига, както го знае приложението
    lookup = {}
    for c in chains:
        lookup[_geo_norm(c)] = c
    for c, aliases in GEO_ALIASES.items():
        for a in aliases:
            lookup[_geo_norm(a)] = c

    query = """
[out:json][timeout:120];
area["ISO3166-1"="BG"][admin_level=2]->.bg;
nwr["shop"~"supermarket|chemist|convenience|discount|greengrocer|variety_store"](area.bg);
out center tags;
"""
    elements = None
    for host in ("https://overpass-api.de/api/interpreter",
                 "https://overpass.kumi.systems/api/interpreter"):
        try:
            r = requests.post(host, data={"data": query}, timeout=180,
                              headers={"User-Agent": "PromoRadar/1.0"})
            if r.ok:
                elements = r.json().get("elements", [])
                break
            note("гео", f"{host}: HTTP {r.status_code}")
        except Exception as e:
            note("гео", f"{host}: {type(e).__name__}: {e}")

    if elements is None:
        prev = _geo_prev()
        if prev:
            OUT["geo"] = prev
            OUT["stats"]["гео"] = (f"Overpass недостъпен — наследени "
                                   f"{sum(len(v) for v in prev.values())} "
                                   f"магазина от предишния фийд")
        else:
            note("гео", "Overpass недостъпен и няма предишен файл")
        return

    by_chain = {}
    for el in elements:
        tags = el.get("tags", {})
        lat = el.get("lat") or el.get("center", {}).get("lat")
        lon = el.get("lon") or el.get("center", {}).get("lon")
        if lat is None or lon is None:
            continue
        chain = None
        for key in ("brand", "name", "operator"):
            v = tags.get(key)
            if v and _geo_norm(v) in lookup:
                chain = lookup[_geo_norm(v)]
                break
        if chain is None:
            continue
        by_chain.setdefault(chain, []).append(
            [round(float(lat), 4), round(float(lon), 4)])

    # дедупликация на съседни точки (двойни записи в OSM) и разумен таван
    for c, pts in by_chain.items():
        seen, ded = set(), []
        for p in pts:
            k = (round(p[0], 3), round(p[1], 3))  # ~110 м клетка
            if k not in seen:
                seen.add(k)
                ded.append(p)
        by_chain[c] = ded[:500]

    if not by_chain:
        note("гео", f"нито един от {len(elements)} OSM обекта не съвпадна с верига")
        return
    OUT["geo"] = by_chain
    OUT["stats"]["гео"] = (f"{sum(len(v) for v in by_chain.values())} магазина "
                           f"в {len(by_chain)} вериги от {len(elements)} OSM обекта")


# ---------------------------------------------------------------------------
# AI етикети — смисълът, който правилата не могат да уловят
# ---------------------------------------------------------------------------
#
# Правилата разбират думи; „BRIO Буркан“ срещу „БОБ БУРКАН“ се различават
# по СМИСЪЛ. Затова веднъж на пускане езиков модел преглежда САМО новите
# имена (кешът в feed/labels.json помни всяко видяно име завинаги) и
# връща категория, марка и тип продукт. Телефонът получава готовите
# етикети — нула AI на устройството. Без ключ (ANTHROPIC_API_KEY в
# GitHub Secrets) стъпката просто се прескача и правилата поемат всичко.

AI_CATS = ("plod meso mlek hlyab bakal napit sladko zamr higi dom teh "
           "dreh pet bebe tut drugo").split()

def _lab_key(name):
    return re.sub(r"\s+", " ", name).strip().lower()[:80]

def _labels_prev():
    base = CFG.get("feed_raw_base", "").rstrip("/")
    if not base:
        return {}
    for br in ("data", "main"):
        try:
            r = requests.get(f"{base}/{br}/feed/labels.json", timeout=30)
            if r.ok:
                return r.json().get("l", {})
        except Exception:
            pass
    return {}

def _ai_batch(names, key, model):
    """Едно повикване: до 80 имена → списък с етикети."""
    lines = "\n".join(f"{i}|{n}" for i, n in enumerate(names))
    prompt = (
        "Ти си класификатор на продукти от български супермаркети. За всеки "
        "ред (номер|име) върни JSON масив с обекти {\"i\":номер, \"c\":категория, "
        "\"b\":марка, \"t\":тип, \"p\":каноничен продукт, \"q\":количество}. "
        "Категорията е ЕДНА от: " + " ".join(AI_CATS) +
        ". Марката е с малки букви, както е в името, или \"\" ако няма. Типът е "
        "кратък (2-4 думи, малки букви) и описва КАКВО Е продуктът, не опаковката: "
        "„празен буркан“ е „буркан за консервиране“ (teh), „БОБ БУРКАН“ е „боб“ "
        "(bakal); паста за зъби с вкус ягода е higi, не плод; цигарите са tut. "
        "Каноничният продукт (p) е ЕДНАКЪВ за всички изписвания на един и същ "
        "продукт: марка + тип + вариант + количество, малки букви, точно в този "
        "ред — „ПРЯСНО МЛЯКО ВЕРЕЯ 3% 1Л“ и „Верея прясно мляко 3.0% 1 l“ дават "
        "еднакво \"верея прясно мляко 3% 1л\"; без марка се пропуска марката. "
        "Количеството (q) е ОБЩОТО в опаковката, само число и единица от г|мл|бр "
        "— „3x100г“ дава \"300 г\", „1,5 л“ дава \"1500 мл\", неизвестно = \"\". "
        "Върни САМО JSON масива, без друг текст.\n\n" + lines)
    r = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={"x-api-key": key, "anthropic-version": "2023-06-01",
                 "content-type": "application/json"},
        json={"model": model, "max_tokens": 6000,
              "messages": [{"role": "user", "content": prompt}]},
        timeout=120)
    r.raise_for_status()
    text = r.json()["content"][0]["text"].strip()
    m = re.search(r"\[.*\]", text, re.S)
    out = {}
    for row in json.loads(m.group(0) if m else text):
        try:
            i = int(row["i"])
            c = row.get("c", "")
            if c in AI_CATS and 0 <= i < len(names):
                out[names[i]] = {"c": c,
                                 "b": str(row.get("b", ""))[:30].lower(),
                                 "t": str(row.get("t", ""))[:40].lower(),
                                 "p": str(row.get("p", ""))[:70].lower(),
                                 "q": str(row.get("q", ""))[:15].lower()}
        except Exception:
            continue
    return out

def enrich_ai():
    import os
    labels = _labels_prev()
    today = datetime.date.today().isoformat()

    all_names = ([o["name"] for o in OUT["offers"]] +
                 [b["product"] for b in OUT["basics"]])
    seen_keys = set()
    fresh = []
    for n in all_names:
        k = _lab_key(n)
        if k in seen_keys:
            continue
        seen_keys.add(k)
        if k in labels:
            labels[k]["l"] = today
        else:
            fresh.append((k, n))

    key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    model = CFG.get("ai_model", "claude-haiku-4-5")
    max_new = int(CFG.get("ai_max_new_per_run", 4000))
    done = 0
    if key and fresh:
        batch_names = [n for _, n in fresh[:max_new]]
        key_of = {n: k for k, n in fresh[:max_new]}
        for i in range(0, len(batch_names), 80):
            chunk = batch_names[i:i + 80]
            try:
                got = _ai_batch(chunk, key, model)
            except Exception as e:
                # една счупена партида не бива да спира останалите —
                # неетикетираните ще дойдат пак при следващото пускане
                note("ai", f"партида {i//80}: {type(e).__name__}: {e}")
                continue
            for n, lab in got.items():
                lab["l"] = today
                labels[key_of[n]] = lab
                done += 1
    elif not key and fresh:
        OUT["stats"]["ai"] = (f"{len(fresh)} нови имена чакат — няма "
                              f"ANTHROPIC_API_KEY в secrets; правилата поемат")

    # Досъбиране: старите етикети нямат каноничен продукт (p) — минават
    # повторно, в рамките на същия бюджет, докато всички го получат.
    backfilled = 0
    if key:
        room = max_new - done
        redo = [k for k, v in labels.items()
                if "p" not in v and v.get("l") == today][:room]
        for i in range(0, len(redo), 80):
            chunk = redo[i:i + 80]
            try:
                got = _ai_batch(chunk, key, model)
            except Exception as e:
                note("ai", f"досъбиране {i//80}: {type(e).__name__}: {e}")
                continue
            for n, lab in got.items():
                lab["l"] = labels.get(n, {}).get("l", today)
                labels[n] = lab
                backfilled += 1
            # AI-ят прегледа целия пакет: имената, за които не върна
            # канонично име (карти, отстъпки, общи текстове), получават
            # празно "p" — прегледано е, няма какво да се измисли, и не
            # се пращат отново при всяко пускане.
            for n in chunk:
                if n in labels and "p" not in labels[n]:
                    labels[n]["p"] = ""

    # чистене: имена, невиждани от 90 дни, си отиват с офертите
    cutoff = (datetime.date.today() - datetime.timedelta(days=90)).isoformat()
    labels = {k: v for k, v in labels.items() if v.get("l", today) >= cutoff}

    OUT["labels"] = labels
    if key:
        # броим само реално чакащите: имена от днешните оферти без "p".
        # Старите без "p", които вече не се появяват, не се пращат никъде
        # и не струват нищо — не влизат в брояча.
        no_p = sum(1 for v in labels.values()
                   if "p" not in v and v.get("l") == today)
        OUT["stats"]["ai"] = (f"{len(labels)} етикета в кеша, {done} нови, "
                              f"{backfilled} досъбрани канонични, "
                              f"{max(0, len(fresh)-done)} чакат нови, "
                              f"{no_p} чакат канонично име")


# ---------------------------------------------------------------------------
# Количествата от AI → цена за единица там, където правилата не стигат
# ---------------------------------------------------------------------------
#
# Regex-ът хваща „500 г“, но не и „3x100г“ или „кашон 12 бр“. AI етикетът
# носи ОБЩОТО количество (q) — тук то се превръща в цена за единица за
# редовете, които още нямат. Така подредбата „за единица“ и сравнението
# на опаковки покриват почти всичко.

_AI_Q_RE = re.compile(r"^(\d+(?:[.,]\d+)?)\s*(г|мл|бр)$")

def _unit_from_q(price, q):
    m = _AI_Q_RE.match(q.strip())
    if not m or price <= 0:
        return None
    v = float(m.group(1).replace(",", "."))
    if v <= 0:
        return None
    unit = m.group(2)
    if unit == "г":
        return round(price / (v / 1000.0), 2), "лв/кг"
    if unit == "мл":
        return round(price / (v / 1000.0), 2), "лв/л"
    return round(price / v, 2), "лв/бр"

def apply_ai_quantities():
    labels = OUT.get("labels") or {}
    if not labels:
        return
    filled = 0
    for o in OUT["offers"]:
        if o.get("unitPrice") is not None:
            continue
        lab = labels.get(_lab_key(o["name"]))
        if not lab or not lab.get("q"):
            continue
        got = _unit_from_q(float(o["price"]), lab["q"])
        if got:
            o["unitPrice"], o["unitLabel"] = got
            filled += 1
    for b in OUT["basics"]:
        if b.get("unitPrice") is not None:
            continue
        lab = labels.get(_lab_key(b["product"]))
        if not lab or not lab.get("q"):
            continue
        got = _unit_from_q(float(b["price"]), lab["q"])
        if got:
            b["unitPrice"], b["unitLabel"] = got
            filled += 1
    if filled:
        OUT["stats"]["ai_количества"] = f"{filled} реда получиха цена за единица от AI"


# ---------------------------------------------------------------------------
# Одит на имената — ранното предупреждение за „сапун при пъпешите“
# ---------------------------------------------------------------------------
#
# Историята: „БОЧКО ПЗ ПЪПЕШ“ и „CH Aroma NE пъпеш“ се показваха като
# подобни на истинския пъпеш, защото класификаторът в приложението не
# познаваше марката. Този одит хваща същия модел ОЩЕ ПРИ ОБХОЖДАНЕТО:
# име с хранителна дума + козметичен белег, но БЕЗ познатите котви
# (Ш-Н, С-Н, ПЗ, шампоан…), отива в статистиката. Така новата марка се
# вижда в лога на workflow-а, преди човек да я види в приложението.

_AUDIT_FOOD = re.compile(
    r"пъпеш|ябълк|ягод|праск|лимон|портокал|банан|кокос|авокадо|череш|"
    r"боровинк|малин|киви|манго|мляко|яйц|месо|пилеш", re.I)
_AUDIT_COSM = re.compile(
    r"\b\d+\s*(мл|ml)\b|арома\b|aroma|парфюм", re.I)
_AUDIT_ANCHOR = re.compile(
    r"ш-н|с-н|\bпз\b|шампоан|сапун|балсам|душ гел|паста за зъби|крем|"
    r"лосион|кърпи|кърпичк|дезодор|бочко|пуфис|колгейт|мицелар|"
    r"сок|нектар|напитк|вода|кафе|чай|бира|вино|йогурт|десерт|"
    # течни ХРАНИ в милилитри и лекарства с познати котви — не са аларма
    r"мляко|оцет|олио|комбуча|сироп|витамин|карнитин|таблетк|капсул|"
    r"сусп|\bмг\b|нурофен|стрепсилс|ангал|препарат|съд|fairy|frosch|"
    r"фея|пюре|бебе|bebelan|шардоне|мерло|совиньон|ликьор|гел|спрей", re.I)

def audit_names():
    sus = []
    for o in OUT["offers"] + [
        {"store": b["chain"], "name": b["product"]} for b in OUT["basics"]
    ]:
        n = o["name"]
        if (_AUDIT_FOOD.search(n) and _AUDIT_COSM.search(n)
                and not _AUDIT_ANCHOR.search(n)):
            sus.append(f'{o["store"]}: {n[:60]}')
    if sus:
        OUT["stats"]["одит_подозрителни"] = sus[:15]
        OUT["stats"]["одит"] = (f"{len(sus)} имена приличат на козметика с "
                                f"хранителна дума — провери дали не трябват "
                                f"нови стеми в Categories.kt")
        _audit_ai(sus[:15])
    else:
        OUT["stats"]["одит"] = "чисто — нито едно подозрително име"


def _audit_ai(sus):
    """Автопилотът: AI преглежда подозрителните и предлага готовите стеми.

    Предложенията отиват само в статистиката — човек ги одобрява, преди
    да влязат в речника. Едно повикване на пускане, стотинка."""
    import os
    key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not key:
        return
    model = CFG.get("ai_model", "claude-haiku-4-5")
    lines = "\n".join(s.split(": ", 1)[-1] for s in sus)
    prompt = (
        "Това са продукти от български магазини, които класификатор по думи "
        "може да сбърка (хранителна дума в нехранителен продукт или обратно). "
        "За всеки ред върни JSON масив с {\"n\":първите 30 знака от името, "
        "\"c\":правилната категория от " + " ".join(AI_CATS) + ", "
        "\"s\":думата-котва от името, която еднозначно издава категорията "
        "(марка, съкращение или тип, с малки букви)}. Само JSON.\n\n" + lines)
    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": key, "anthropic-version": "2023-06-01",
                     "content-type": "application/json"},
            json={"model": model, "max_tokens": 1500,
                  "messages": [{"role": "user", "content": prompt}]},
            timeout=90)
        r.raise_for_status()
        text = r.json()["content"][0]["text"].strip()
        m = re.search(r"\[.*\]", text, re.S)
        rows = json.loads(m.group(0) if m else text)
        OUT["stats"]["одит_ai_предложения"] = [
            f'{str(x.get("n",""))[:30]} → {x.get("c","")} (стем: „{x.get("s","")}“)'
            for x in rows if x.get("c") in AI_CATS][:15]
    except Exception as e:
        note("одит-ai", f"{type(e).__name__}: {e}")


def build_digest():
    """Дневният дайджест: 2-3 човешки изречения за най-добрите сделки.

    Отива във feed/digest.json — приложението и уеб версията го показват
    като бележка на деня. Едно повикване на пускане."""
    import os
    key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not key or not OUT["offers"]:
        return
    model = CFG.get("ai_model", "claude-haiku-4-5")
    top = sorted(
        OUT["offers"],
        key=lambda o: -(o.get("pct") or (
            int((1 - o["price"] / o["old"]) * 100) if o.get("old") else 0))
    )[:8]
    lines = "\n".join(
        f'{o["store"]}: {o["name"][:55]} — {o["price"]:.2f} лв'
        + (f' (беше {o["old"]:.2f})' if o.get("old") else "")
        for o in top)
    prompt = (
        "Ти пишеш дневната бележка на приложение за промоции. От тези "
        "най-големи намаления днес напиши 2-3 кратки изречения на български "
        "— кое си струва и защо, делово и без възклицания, без емоджита, "
        "цените в евро (лв ÷ 1.95583, закръглени до стотинка). "
        "Върни САМО текста.\n\n" + lines)
    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": key, "anthropic-version": "2023-06-01",
                     "content-type": "application/json"},
            json={"model": model, "max_tokens": 400,
                  "messages": [{"role": "user", "content": prompt}]},
            timeout=90)
        r.raise_for_status()
        text = r.json()["content"][0]["text"].strip()[:500]
        if text:
            OUT["digest"] = text
            OUT["stats"]["дайджест"] = text[:80] + "…"
    except Exception as e:
        note("дайджест", f"{type(e).__name__}: {e}")


def write_output():
    """Пише разделен фийд + пълния файл за съвместимост.

    feed/index.json      малък: кои вериги има, колко продукта, топ офертите
    feed/offers-<x>.json офертите на една верига — тегли се при нужда
    feed/search.json     компактен индекс за търсене
    feed/basics.json     официалните цени — тегли се само при „къде е най-евтино“
    data.json            пълният файл (както досега), за съвместимост
    """
    import os
    sizes = {}

    # --- по вериги ---
    stores = []
    by_store = {}
    for o in OUT["offers"]:
        by_store.setdefault(o["store"], []).append(o)
    for name, lst in sorted(by_store.items(), key=lambda kv: -len(kv[1])):
        sl = slug(name)
        sizes[f"offers-{sl}"] = _dump(f"feed/offers-{sl}.json",
                                      {"store": name, "updated": OUT["updated"],
                                       "offers": lst})
        stores.append({"name": name, "slug": sl, "count": len(lst)})

    # --- официалните цени (тежки, но нужни само за кошницата) ---
    chains = sorted({b["chain"] for b in OUT["basics"]})
    sizes["basics"] = _dump("feed/basics.json",
                            {"updated": OUT["updated"],
                             "date": OUT.get("basics_date", ""),
                             "chains": chains,
                             "basics": OUT["basics"]})

    # --- подадените цени (малък файл, но носи магазините, които никой не обхожда) ---
    if OUT["community"]:
        sizes["community"] = _dump("feed/community.json",
                                   {"updated": OUT["updated"],
                                    "prices": OUT["community"]})

    # --- историята на цените (тегли се при отваряне на продукт) ---
    if OUT["history"]:
        sizes["history"] = _dump("feed/history.json",
                                 {"updated": OUT["updated"],
                                  "h": OUT["history"]})

    # --- дневният дайджест (бележката на деня) ---
    if OUT.get("digest"):
        sizes["digest"] = _dump("feed/digest.json",
                                {"updated": OUT["updated"],
                                 "text": OUT["digest"]})

    # --- AI етикетите (категория/марка/тип за всяко име) ---
    if OUT.get("labels"):
        sizes["labels"] = _dump("feed/labels.json",
                                {"updated": OUT["updated"],
                                 "l": OUT["labels"]})

    # --- локациите на магазините (за „вериги наблизо“) ---
    if OUT.get("geo"):
        sizes["stores-geo"] = _dump("feed/stores-geo.json",
                                    {"updated": OUT["updated"],
                                     "chains": OUT["geo"]})

    # --- компактен индекс за търсене (кратки ключове = малък файл) ---
    idx = []
    for o in OUT["offers"]:
        idx.append({"s": o["store"], "n": o["name"], "p": o["price"],
                    "o": o.get("old"), "u": o.get("unitPrice"),
                    "l": o.get("unitLabel", ""), "f": 0})
    for b in OUT["basics"]:
        idx.append({"s": b["chain"], "n": b["product"], "p": b["price"],
                    "o": None, "u": b.get("unitPrice"),
                    "l": b.get("unitLabel", ""), "f": 1})
    for c in OUT["community"]:
        idx.append({"s": c["chain"], "n": c["product"], "p": c["price"],
                    "o": None, "u": c.get("unitPrice"),
                    "l": c.get("unitLabel", ""), "f": 2})
    sizes["search"] = _dump("feed/search.json",
                            {"updated": OUT["updated"], "items": idx})

    # --- индексът: това е единственото, което се тегли при всяко отваряне ---
    top = sorted(
        OUT["offers"],
        key=lambda o: -(o.get("pct") or (
            int((1 - o["price"] / o["old"]) * 100) if o.get("old") else 0))
    )[:200]
    index = {
        "updated": OUT["updated"], "version": 3,
        "stores": stores, "chains": chains,
        "counts": {"offers": len(OUT["offers"]), "basics": len(OUT["basics"]),
                   "ebag": len(OUT["ebag"])},
        "ebag": OUT["ebag"],
        "top": top,
        "errors": OUT["errors"][:40],
        "stats": OUT["stats"],
    }
    sizes["index"] = _dump("feed/index.json", index)

    # --- пълният файл (както досега) ---
    sizes["data"] = _dump("data.json", OUT)

    OUT["stats"]["размери"] = {k: f"{v/1024:.0f} KB" for k, v in sizes.items()}
    OUT["stats"]["тегли се при отваряне"] = f"{sizes['index']/1024:.0f} KB"
    if sizes["data"] > 6_000_000:
        note("размер", f"пълният data.json е {sizes['data']/1024/1024:.1f} MB — "
                       f"новото приложение тегли само index.json "
                       f"({sizes['index']/1024:.0f} KB), но ако ползваш стара "
                       f"версия, намали max_per_store")
    # презапис със сметките за размерите вътре
    _dump("feed/index.json", {**index, "stats": OUT["stats"]})
    _dump("data.json", OUT)


# ---------------------------------------------------------------------------
def main():
    ctx = browser = None
    try:
        from playwright.sync_api import sync_playwright
        pw = sync_playwright().start()
        browser = pw.chromium.launch()
        ctx = browser.new_context(viewport={"width": 1280, "height": 2400},
                                  locale="bg-BG", user_agent=UA["User-Agent"])
    except Exception as e:
        log_err("playwright/start", e)

    try:
        scrape_ebag(ctx)
        scrape_dom_offers(ctx)
    finally:
        if browser:
            try:
                browser.close()
            except Exception:
                pass

    scrape_basics()
    scrape_community()
    scrape_reports()
    build_history()
    build_geo()
    enrich_ai()
    apply_ai_quantities()
    audit_names()
    build_digest()

    OUT["stats"]["време общо"] = f"{elapsed_min():.1f} мин"
    OUT["stats"]["offers_total"] = len(OUT["offers"])
    OUT["stats"]["with_unit_price"] = sum(
        1 for o in OUT["offers"] + OUT["basics"] if "unitPrice" in o)

    write_output()

    print(f"ebag={len(OUT['ebag'])} offers={len(OUT['offers'])} "
          f"basics={len(OUT['basics'])} errors={len(OUT['errors'])}")
    for k, v in OUT["stats"].items():
        print(f"  {k}: {v}")
    for e in OUT["errors"][:20]:
        print(f"  ! {e}")


if __name__ == "__main__":
    main()
